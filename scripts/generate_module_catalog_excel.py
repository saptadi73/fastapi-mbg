from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("docs/module_catalog_erp_mbg_penawaran.xlsx")


MODULES = [
    {
        "group": "Platform & Security",
        "module": "identity",
        "offering_name": "Identity & Access Management",
        "api_prefix": "/api/v1/identity",
        "description": "Mengelola login, profil pengguna, role, dan perpindahan context SPPG aktif.",
        "key_features": "Login, profil user, switch SPPG aktif, manajemen user, akses user per SPPG.",
        "primary_users": "Semua user, tenant admin, super admin",
        "sales_value": "Menjaga keamanan akses, pemisahan kewenangan, dan kesiapan multi-user operasional.",
        "status": "Aktif di backend",
    },
    {
        "group": "Platform & Security",
        "module": "tenant",
        "offering_name": "Tenant Management",
        "api_prefix": "/api/v1/tenants",
        "description": "Mengelola master tenant untuk model bisnis multi-tenant.",
        "key_features": "List tenant, detail tenant, pembuatan tenant.",
        "primary_users": "Super admin",
        "sales_value": "Mendukung model SaaS atau grup operasional dengan pemisahan data antar organisasi.",
        "status": "Aktif di backend",
    },
    {
        "group": "Platform & Security",
        "module": "health",
        "offering_name": "Health Monitoring",
        "api_prefix": "/health",
        "description": "Menyediakan endpoint untuk memeriksa status aplikasi dan database.",
        "key_features": "Live check, readiness check, database health.",
        "primary_users": "Tim IT, DevOps",
        "sales_value": "Membantu monitoring uptime dan kesiapan sistem untuk operasional harian.",
        "status": "Aktif di backend",
    },
    {
        "group": "Master Data",
        "module": "sppg",
        "offering_name": "SPPG / Dapur Management",
        "api_prefix": "/api/v1/sppg",
        "description": "Mengelola profil dapur/SPPG, lokasi, dan konteks operasional utama.",
        "key_features": "List SPPG, detail SPPG, pembuatan SPPG.",
        "primary_users": "Tenant admin, super admin",
        "sales_value": "Menjadi dasar pengelolaan cabang dapur, cakupan layanan, dan segmentasi operasional.",
        "status": "Aktif di backend",
    },
    {
        "group": "Master Data",
        "module": "geography",
        "offering_name": "School & Location Master",
        "api_prefix": "/api/v1/geography",
        "description": "Mengelola master sekolah dan lokasi tujuan layanan.",
        "key_features": "List sekolah, detail sekolah, pembuatan sekolah.",
        "primary_users": "Operations manager, tenant admin",
        "sales_value": "Membantu pemetaan area layanan dan target distribusi secara terstruktur.",
        "status": "Aktif di backend",
    },
    {
        "group": "Master Data",
        "module": "beneficiary",
        "offering_name": "Beneficiary Management",
        "api_prefix": "/api/v1/beneficiaries",
        "description": "Mengelola data penerima manfaat per sekolah atau unit layanan.",
        "key_features": "List penerima manfaat, detail, pembuatan data beneficiary.",
        "primary_users": "Operations manager, tenant admin",
        "sales_value": "Mendukung akurasi target distribusi dan pelaporan cakupan manfaat program.",
        "status": "Aktif di backend",
    },
    {
        "group": "Master Data",
        "module": "uom",
        "offering_name": "Unit of Measure Management",
        "api_prefix": "/api/v1/uoms",
        "description": "Mengelola satuan ukur yang dipakai dalam stok, resep, dan pembelian.",
        "key_features": "Master UoM dan pemakaian satuan standar.",
        "primary_users": "Tenant admin",
        "sales_value": "Mengurangi kesalahan konversi kuantitas pada proses pengadaan, stok, dan produksi.",
        "status": "Aktif di backend",
    },
    {
        "group": "Master Data",
        "module": "product",
        "offering_name": "Product & Material Master",
        "api_prefix": "/api/v1/products",
        "description": "Mengelola master bahan, produk, dan item yang dipakai lintas modul.",
        "key_features": "List produk, detail produk, pembuatan produk, standard cost.",
        "primary_users": "Tenant admin, operations manager, procurement officer",
        "sales_value": "Menjadi sumber data utama untuk resep, stok, procurement, dan costing.",
        "status": "Aktif di backend",
    },
    {
        "group": "Master Data",
        "module": "recipe",
        "offering_name": "Recipe Management",
        "api_prefix": "/api/v1/recipes",
        "description": "Mengelola resep dan komposisi bahan untuk perencanaan menu dan produksi.",
        "key_features": "Master resep, line resep, versioning resep dasar.",
        "primary_users": "Tenant admin, operations manager",
        "sales_value": "Menstandarkan formulasi menu dan perhitungan kebutuhan bahan baku.",
        "status": "Aktif di backend",
    },
    {
        "group": "Program & Planning",
        "module": "program",
        "offering_name": "Program Management",
        "api_prefix": "/api/v1",
        "description": "Mengelola program, periode program, serta assignment tenant dan SPPG ke program.",
        "key_features": "List program, detail, pembuatan program, assignment tenant/SPPG, period management.",
        "primary_users": "Tenant admin, super admin",
        "sales_value": "Mendukung pengelolaan program bantuan atau inisiatif operasional yang memiliki periode dan scope jelas.",
        "status": "Aktif di backend",
    },
    {
        "group": "Program & Planning",
        "module": "meal_plan",
        "offering_name": "Meal Planning",
        "api_prefix": "/api/v1/meal-plans",
        "description": "Mengelola rencana menu, jumlah porsi, dan estimasi biaya per periode layanan.",
        "key_features": "Meal plan list/detail, submit, approve, reserve materials, cost preview.",
        "primary_users": "Operations manager, tenant admin, finance manager",
        "sales_value": "Membantu kontrol rencana produksi dan biaya sebelum operasional dijalankan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Supply Chain",
        "module": "procurement",
        "offering_name": "Procurement Management",
        "api_prefix": "/api/v1/procurement",
        "description": "Mengelola proses pembelian dari purchase request sampai supplier payment.",
        "key_features": "Supplier, PR, PO, goods receipt, supplier invoice, supplier payment.",
        "primary_users": "Procurement officer, finance manager, operations manager",
        "sales_value": "Mempercepat siklus pengadaan sambil menjaga jejak biaya dan kelengkapan dokumen.",
        "status": "Aktif di backend",
    },
    {
        "group": "Supply Chain",
        "module": "inventory",
        "offering_name": "Inventory & Warehouse",
        "api_prefix": "/api/v1/inventory",
        "description": "Mengelola gudang, lokasi stok, batch, transaksi inventori, dan saldo stok.",
        "key_features": "Warehouse, stock location, batch, inventory transaction, balance, expiry alert, FEFO preview.",
        "primary_users": "Warehouse operator, operations manager, procurement officer",
        "sales_value": "Meningkatkan visibilitas stok dan menekan risiko kekurangan atau kadaluarsa bahan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Operations",
        "module": "production",
        "offering_name": "Production Management",
        "api_prefix": "/api/v1/production-orders",
        "description": "Mengelola order produksi, penyelesaian produksi, dan pencatatan biaya aktual material.",
        "key_features": "Production order, completion, material consumption, cost sheet.",
        "primary_users": "Operations manager, finance manager",
        "sales_value": "Menghubungkan rencana menu dengan realisasi produksi dan biaya aktual dapur.",
        "status": "Aktif di backend",
    },
    {
        "group": "Operations",
        "module": "quality",
        "offering_name": "Quality Control",
        "api_prefix": "/api/v1",
        "description": "Mengelola inspeksi mutu untuk memastikan hasil produksi memenuhi standar.",
        "key_features": "QC inspection list/detail/create dan line inspeksi.",
        "primary_users": "Quality officer, operations manager",
        "sales_value": "Menjaga kualitas layanan dan mendukung kepatuhan proses mutu.",
        "status": "Aktif di backend",
    },
    {
        "group": "Operations",
        "module": "delivery",
        "offering_name": "Delivery & Distribution",
        "api_prefix": "/api/v1/delivery-orders",
        "description": "Mengelola pengiriman, route, proof of delivery, dan incident distribusi.",
        "key_features": "Delivery order, route, delivery proof, delivery incident.",
        "primary_users": "Delivery officer, operations manager, quality officer",
        "sales_value": "Meningkatkan kontrol pengiriman sampai bukti serah terima di titik tujuan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Operations",
        "module": "fleet",
        "offering_name": "Fleet & Vehicle Management",
        "api_prefix": "/api/v1/fleet",
        "description": "Mengelola armada, driver, assignment kendaraan, maintenance, dan tracking lokasi.",
        "key_features": "Vehicle type, vehicle, driver, assignment, maintenance, live GPS tracking.",
        "primary_users": "Operations manager, tenant admin, delivery officer",
        "sales_value": "Memberi visibilitas penuh pada armada distribusi dan utilisasi kendaraan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Operations",
        "module": "workforce",
        "offering_name": "Workforce Management",
        "api_prefix": "/api/v1/workforce",
        "description": "Mengelola tenaga kerja, shift, attendance, timesheet, dan labor cost.",
        "key_features": "Position, employee, assignment, shift, attendance, timesheet, labor cost.",
        "primary_users": "Tenant admin, operations manager, finance manager",
        "sales_value": "Membantu kontrol produktivitas SDM dan biaya tenaga kerja per dapur.",
        "status": "Aktif di backend",
    },
    {
        "group": "Operations",
        "module": "gis",
        "offering_name": "GIS & Service Area",
        "api_prefix": "/api/v1/gis",
        "description": "Mengelola analitik spasial, area layanan, dan validasi assignment secara geografis.",
        "key_features": "Service area, coverage validation, spatial assignment validation.",
        "primary_users": "Operations manager, planner",
        "sales_value": "Membantu ekspansi layanan dan validasi jangkauan operasional berbasis peta.",
        "status": "Aktif di backend",
    },
    {
        "group": "Finance",
        "module": "budget",
        "offering_name": "Budget Management",
        "api_prefix": "/api/v1",
        "description": "Mengelola budget, submit/approve budget, dan monitoring availability.",
        "key_features": "Budget header/line, approval, budget availability, reserved/committed/actual.",
        "primary_users": "Finance manager, tenant admin",
        "sales_value": "Membantu pengendalian anggaran dan menjaga disiplin penggunaan dana.",
        "status": "Aktif di backend",
    },
    {
        "group": "Finance",
        "module": "accounting",
        "offering_name": "Accounting & Journal",
        "api_prefix": "/api/v1",
        "description": "Mengelola chart of accounts dan jurnal manual maupun operasional.",
        "key_features": "Accounts, journal entries, posting jurnal, source document traceability.",
        "primary_users": "Finance manager, tenant admin",
        "sales_value": "Menjadi fondasi pembukuan dan sumber data utama laporan keuangan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Finance",
        "module": "costing",
        "offering_name": "Costing & Variance Analysis",
        "api_prefix": "/api/v1",
        "description": "Mengelola kebijakan costing dan analisis biaya produksi aktual versus budget.",
        "key_features": "Cost policy, production cost summary, variance calculation.",
        "primary_users": "Finance manager, operations manager",
        "sales_value": "Memberi insight profitabilitas dan pengendalian biaya per produksi atau dapur.",
        "status": "Aktif di backend",
    },
    {
        "group": "Finance",
        "module": "government_claim",
        "offering_name": "Government Claim Management",
        "api_prefix": "/api/v1/government-claims",
        "description": "Mengelola klaim ke pemerintah dari persiapan, verifikasi, adjustment, sampai pembayaran.",
        "key_features": "Claim, evidence, verification, adjustment, payment, aging source data.",
        "primary_users": "Finance manager, tenant admin",
        "sales_value": "Mempercepat proses pencairan klaim dan memudahkan monitoring piutang pemerintah.",
        "status": "Aktif di backend",
    },
    {
        "group": "Finance",
        "module": "funding",
        "offering_name": "Funding & Investor Position",
        "api_prefix": "/api/v1/funding",
        "description": "Mengelola sumber pendanaan, agreement, disbursement, dan repayment investor.",
        "key_features": "Funding source, funding agreement, disbursement, repayment, funding position.",
        "primary_users": "Finance manager, tenant admin",
        "sales_value": "Membantu perusahaan mengelola modal kerja dan transparansi kewajiban pendanaan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Finance",
        "module": "asset",
        "offering_name": "Asset Management",
        "api_prefix": "/api/v1/assets",
        "description": "Mengelola kategori asset, master asset, assignment asset, dan depresiasi.",
        "key_features": "Asset category, asset register, asset assignment, asset depreciation.",
        "primary_users": "Finance manager, operations manager",
        "sales_value": "Membantu kontrol aset operasional dan pencatatan depresiasi untuk laporan keuangan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Engagement & Governance",
        "module": "feedback",
        "offering_name": "Feedback & Complaint Management",
        "api_prefix": "/api/v1/feedback",
        "description": "Mengelola feedback layanan, complaint, dan penilaian kualitas layanan.",
        "key_features": "Feedback submission, complaint, service quality score.",
        "primary_users": "Operations manager, quality officer",
        "sales_value": "Membantu meningkatkan kepuasan penerima manfaat dan kualitas layanan lapangan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Engagement & Governance",
        "module": "document",
        "offering_name": "Document Management",
        "api_prefix": "/api/v1",
        "description": "Mengelola metadata dokumen, versi dokumen, dan relasi dokumen ke transaksi bisnis.",
        "key_features": "Document, versioning, link ke entity bisnis.",
        "primary_users": "Tenant admin, operations manager, quality officer, finance manager",
        "sales_value": "Memudahkan pengarsipan dokumen dan keterlacakan bukti operasional maupun keuangan.",
        "status": "Aktif di backend",
    },
    {
        "group": "Engagement & Governance",
        "module": "workflow",
        "offering_name": "Workflow & Approval Engine",
        "api_prefix": "/api/v1",
        "description": "Mengelola workflow generic untuk approval dokumen lintas modul.",
        "key_features": "Workflow definition, transition, document workflow viewer, approval flow.",
        "primary_users": "Tenant admin, super admin, approver lintas fungsi",
        "sales_value": "Meningkatkan kontrol tata kelola dan disiplin approval proses bisnis.",
        "status": "Aktif di backend",
    },
    {
        "group": "Engagement & Governance",
        "module": "notification",
        "offering_name": "Notification Center",
        "api_prefix": "/api/v1/notifications",
        "description": "Mengelola inbox notifikasi, delivery status, dan antrian notifikasi operasional.",
        "key_features": "Inbox notification, mark as read, enqueue notification.",
        "primary_users": "Semua user",
        "sales_value": "Membantu pengguna merespons approval, alert, dan aktivitas penting secara cepat.",
        "status": "Aktif di backend",
    },
    {
        "group": "Engagement & Governance",
        "module": "audit",
        "offering_name": "Audit Trail",
        "api_prefix": "/api/v1",
        "description": "Mencatat jejak audit perubahan dan aktivitas penting lintas modul.",
        "key_features": "Audit event list, detail audit event, module traceability.",
        "primary_users": "Tenant admin, super admin, auditor internal",
        "sales_value": "Mendukung kepatuhan, investigasi, dan kontrol internal atas transaksi penting.",
        "status": "Aktif di backend",
    },
    {
        "group": "Engagement & Governance",
        "module": "reporting",
        "offering_name": "Executive Reporting & Dashboard",
        "api_prefix": "/api/v1",
        "description": "Menyediakan dashboard operasional dan laporan finance seperti cash flow, laba rugi, dan neraca.",
        "key_features": "Tenant dashboard, SPPG dashboard, finance dashboard, cash flow, profit loss, balance sheet, ROI, aging.",
        "primary_users": "Management, finance manager, operations manager",
        "sales_value": "Menyediakan insight manajerial cepat untuk pengambilan keputusan dan pelaporan mitra.",
        "status": "Aktif di backend",
    },
    {
        "group": "Platform & Analytics",
        "module": "ai",
        "offering_name": "AI Assistant & Analytics",
        "api_prefix": "/api/v1/ai",
        "description": "Menyediakan fitur AI seperti forecast, recommendation, daily summary, dan NL2SQL.",
        "key_features": "AI forecasts, recommendations, daily summaries, natural language to SQL.",
        "primary_users": "Tenant admin, operations manager, finance manager",
        "sales_value": "Menambah nilai inovasi dengan insight prediktif dan eksplorasi data berbasis bahasa natural.",
        "status": "Aktif di backend",
    },
    {
        "group": "Platform & Analytics",
        "module": "integration",
        "offering_name": "Integration Hub",
        "api_prefix": "/api/v1",
        "description": "Mengelola external system, credential metadata, webhook, data mapping, dan sync job.",
        "key_features": "External system registry, webhook subscription, data mapping, sync jobs.",
        "primary_users": "Tenant admin, technical integration team",
        "sales_value": "Membuka jalan integrasi dengan sistem eksternal seperti ERP, BI, atau layanan pemerintah.",
        "status": "Aktif di backend",
    },
    {
        "group": "Platform & Analytics",
        "module": "platform_ops",
        "offering_name": "Platform Operations",
        "api_prefix": "/api/v1/platform",
        "description": "Mengelola background job, outbox event, dan read model operasional platform.",
        "key_features": "Background jobs, outbox events, refresh summary tables/materialized views.",
        "primary_users": "Super admin, tenant admin, tim IT",
        "sales_value": "Menjaga stabilitas platform dan mendukung otomasi proses backend skala besar.",
        "status": "Aktif di backend",
    },
]


GROUP_SUMMARY = [
    ("Platform & Security", "Fondasi aplikasi, keamanan akses, dan pengelolaan multi-tenant."),
    ("Master Data", "Sumber data utama yang dipakai lintas proses operasional dan pelaporan."),
    ("Program & Planning", "Perencanaan menu, cakupan program, dan penjadwalan kegiatan."),
    ("Supply Chain", "Pengadaan, stok, dan aliran bahan sampai siap dipakai produksi."),
    ("Operations", "Pelaksanaan harian dapur, distribusi, kualitas, armada, dan SDM."),
    ("Finance", "Kontrol anggaran, akuntansi, costing, klaim, pendanaan, dan aset."),
    ("Engagement & Governance", "Dokumen, feedback, workflow, notifikasi, audit, dan laporan eksekutif."),
    ("Platform & Analytics", "Integrasi, AI, dan otomasi operasional platform."),
]


def autosize_columns(worksheet) -> None:
    for index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 14), 48)


def build_workbook() -> Workbook:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Ringkasan"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    overview["A1"] = "Katalog Modul ERP MBG"
    overview["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    overview["A1"].fill = title_fill
    overview.merge_cells("A1:D1")
    overview["A2"] = "Tanggal Pembuatan"
    overview["B2"] = date.today().isoformat()
    overview["A3"] = "Jumlah Modul Aktif"
    overview["B3"] = len(MODULES)
    overview["A5"] = "Kelompok Solusi"
    overview["B5"] = "Penjelasan"
    overview["A5"].font = overview["B5"].font = bold_font
    overview["A5"].fill = overview["B5"].fill = header_fill

    for row_index, (group_name, description) in enumerate(GROUP_SUMMARY, start=6):
        overview.cell(row=row_index, column=1, value=group_name)
        overview.cell(row=row_index, column=2, value=description)

    overview["A16"] = "Catatan Proposal"
    overview["A16"].font = bold_font
    overview["A17"] = (
        "Dokumen ini merangkum modul backend yang sudah terpasang pada aplikasi ERP MBG "
        "dan dapat dipakai sebagai dasar penyusunan proposal solusi kepada mitra atau pelanggan."
    )
    overview["A17"].alignment = wrap_top
    overview.merge_cells("A17:D18")

    executive = workbook.create_sheet("Executive Summary", 1)
    executive["A1"] = "Executive Summary ERP MBG"
    executive["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    executive["A1"].fill = title_fill
    executive.merge_cells("A1:F1")
    executive["A3"] = "Nilai Utama Solusi"
    executive["A3"].font = bold_font
    executive["A4"] = "1. Digitalisasi operasional dapur dari perencanaan sampai distribusi."
    executive["A5"] = "2. Kontrol biaya, anggaran, dan laporan keuangan untuk pengambilan keputusan."
    executive["A6"] = "3. Transparansi proses melalui audit trail, workflow, notifikasi, dan reporting."
    executive["A7"] = "4. Kesiapan ekspansi multi-tenant, multi-SPPG, dan integrasi eksternal."

    executive["A9"] = "Sorotan Modul untuk Proposal"
    executive["A9"].font = bold_font
    executive["A10"] = "Operasional Dapur"
    executive["B10"] = "Meal Planning, Procurement, Inventory, Production, Delivery, Fleet, Workforce"
    executive["A11"] = "Kontrol Finance"
    executive["B11"] = "Budget, Accounting, Costing, Government Claim, Funding, Asset, Reporting"
    executive["A12"] = "Governance"
    executive["B12"] = "Workflow, Document, Notification, Audit"
    executive["A13"] = "Analitik & Inovasi"
    executive["B13"] = "Reporting Dashboard, AI, GIS, Integration Hub"
    for row in range(10, 14):
        executive.cell(row=row, column=1).font = bold_font

    executive["A15"] = "Cocok untuk"
    executive["A15"].font = bold_font
    executive["A16"] = "Penyelenggara dapur MBG, operator multi-cabang, yayasan, koperasi, perusahaan mitra, dan pengelola distribusi pangan terstruktur."
    executive["A16"].alignment = wrap_top
    executive.merge_cells("A16:F17")

    packages = workbook.create_sheet("Paket Penawaran", 2)
    package_headers = ["Paket", "Fokus", "Modul Utama", "Cocok Untuk", "Catatan Penawaran"]
    for col, header in enumerate(package_headers, start=1):
        cell = packages.cell(row=1, column=col, value=header)
        cell.font = white_font
        cell.fill = title_fill
        cell.alignment = wrap_top

    package_rows = [
        (
            "Starter Operations",
            "Digitalisasi operasional inti dapur",
            "Identity, Tenant, SPPG, Geography, Beneficiary, Product, Recipe, Meal Plan, Procurement, Inventory, Production, Delivery",
            "Calon pelanggan yang ingin memulai dari alur perencanaan, pengadaan, produksi, dan distribusi",
            "Paket awal untuk go-live operasional dengan proses inti paling cepat dirasakan manfaatnya.",
        ),
        (
            "Operations + Finance Control",
            "Kontrol biaya dan tata kelola keuangan",
            "Semua modul Starter + Budget, Accounting, Costing, Government Claim, Funding, Asset, Reporting",
            "Organisasi yang membutuhkan transparansi biaya, laporan keuangan, dan pengendalian anggaran",
            "Paket paling kuat untuk proposal ke mitra yang menuntut akuntabilitas finansial dan dashboard manajemen.",
        ),
        (
            "Enterprise Integrated",
            "Skala besar, governance, dan integrasi",
            "Semua modul Operations + Finance + Workflow, Document, Notification, Audit, GIS, AI, Integration, Platform Ops",
            "Operator multi-tenant, multi-SPPG, atau institusi yang ingin otomasi, integrasi, dan analitik lebih lanjut",
            "Paket strategis untuk transformasi end-to-end dan kesiapan integrasi dengan sistem eksternal.",
        ),
    ]
    for row_index, row_values in enumerate(package_rows, start=2):
        for col, value in enumerate(row_values, start=1):
            packages.cell(row=row_index, column=col, value=value)
            packages.cell(row=row_index, column=col).alignment = wrap_top

    modules_sheet = workbook.create_sheet("Daftar Modul")
    headers = [
        "No",
        "Kelompok",
        "Modul Teknis",
        "Nama Modul Penawaran",
        "Prefix API",
        "Penjelasan",
        "Fitur Utama",
        "Pengguna Utama",
        "Nilai untuk Mitra/Pelanggan",
        "Status",
    ]
    for col, header in enumerate(headers, start=1):
        cell = modules_sheet.cell(row=1, column=col, value=header)
        cell.font = white_font
        cell.fill = title_fill
        cell.alignment = wrap_top

    for index, module in enumerate(MODULES, start=2):
        row = index
        modules_sheet.cell(row=row, column=1, value=index - 1)
        modules_sheet.cell(row=row, column=2, value=module["group"])
        modules_sheet.cell(row=row, column=3, value=module["module"])
        modules_sheet.cell(row=row, column=4, value=module["offering_name"])
        modules_sheet.cell(row=row, column=5, value=module["api_prefix"])
        modules_sheet.cell(row=row, column=6, value=module["description"])
        modules_sheet.cell(row=row, column=7, value=module["key_features"])
        modules_sheet.cell(row=row, column=8, value=module["primary_users"])
        modules_sheet.cell(row=row, column=9, value=module["sales_value"])
        modules_sheet.cell(row=row, column=10, value=module["status"])
        for col in range(1, 11):
            modules_sheet.cell(row=row, column=col).alignment = wrap_top

    grouped_sheet = workbook.create_sheet("Kelompok Solusi")
    grouped_headers = ["Kelompok", "Jumlah Modul", "Contoh Modul", "Arah Penawaran"]
    for col, header in enumerate(grouped_headers, start=1):
        cell = grouped_sheet.cell(row=1, column=col, value=header)
        cell.font = white_font
        cell.fill = title_fill
        cell.alignment = wrap_top

    row = 2
    for group_name, group_description in GROUP_SUMMARY:
        modules_in_group = [item for item in MODULES if item["group"] == group_name]
        grouped_sheet.cell(row=row, column=1, value=group_name)
        grouped_sheet.cell(row=row, column=2, value=len(modules_in_group))
        grouped_sheet.cell(
            row=row,
            column=3,
            value=", ".join(item["offering_name"] for item in modules_in_group[:4]),
        )
        grouped_sheet.cell(row=row, column=4, value=group_description)
        for col in range(1, 5):
            grouped_sheet.cell(row=row, column=col).alignment = wrap_top
        row += 1

    for sheet in (overview, executive, packages, modules_sheet, grouped_sheet):
        autosize_columns(sheet)
        if sheet.title == "Ringkasan":
            sheet.freeze_panes = "A6"
        else:
            sheet.freeze_panes = "A2"

    return workbook


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    print(f"Excel katalog modul berhasil dibuat: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
